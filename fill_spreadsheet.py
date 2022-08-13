"""
work in progress
need to replace "from testlab_device_list import device_list as devices" in
imports
and
    for a_device in devices:
        my_proc = Process(target=ip, args=(a_device, return_dict))
in main()
"""
from datetime import datetime
from multiprocessing import Manager, Process
import openpyxl
from openpyxl.styles import Font
from netmiko import ConnectHandler
# from testlab_device_list import device_list as devices


def ip(a_device, return_dict):
    net_connect = ConnectHandler(**a_device)
    output = net_connect.send_command("sh ver", use_textfsm=True)
    output = output[0]
    output0 = output.get("hostname")
    output1 = output.get("version")
    output2 = str(output.get("serial"))
    output2 = (
        output2.replace("'", "").replace("[", "").replace("]", "")
    )  # to remove weird formatting
    output3 = str(output.get("hardware"))
    output3 = (
        output3.replace("'", "").replace("[", "").replace("]", "")
    )  # to remove weird formatting
    output4 = output.get("uptime")
    output5 = net_connect.send_command("sh run int vlan 10 | inc ip")
    output5 = output5.replace("255.255.255.0", "").replace("ip address", "")

    return_dict[str(a_device)] = [output5, output0, output1, output2, output3, output4]


def add_to_wb(return_dict):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("VLAN10")
    sheet["A1"] = "IP Address"
    sheet["B1"] = "Hostname"
    sheet["C1"] = "IOS Version"
    sheet["D1"] = "SN:"
    sheet["E1"] = "Model"
    sheet["F1"] = "Uptime"
    for col in ["A", "B", "C", "D", "E"]:
        sheet.column_dimensions[col].width = 25
    sheet.column_dimensions["F"].width = 45
    sheet.row_dimensions[1].height = 30
    dmx = sheet.row_dimensions[1]
    dmx.font = Font(size=16, bold=True)
    start_row = 3
    for key, device_data in return_dict.items():
        start_column = 1
        for item in device_data:
            output6 = device_data[0].split(".")
            start_row = (int(output6[3])) + 1  # using 4th octet of IP to determine row
            sheet.cell(row=start_row, column=start_column).value = item
            start_column += 1  # start_row would print down/prints first item in list then +1
    wb.save("vlan10_devices.xlsx")
    wb.close()


def main():
    start_time = datetime.now()
    procs = []
    manager = Manager()
    return_dict = manager.dict()
    for a_device in devices:
        my_proc = Process(target=ip, args=(a_device, return_dict))

        my_proc.start()
        procs.append(my_proc)

    for a_proc in procs:
        a_proc.join()

    add_to_wb(return_dict)

    print(datetime.now() - start_time)


if __name__ == "__main__":
    main()
