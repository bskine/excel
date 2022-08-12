from openpyxl import load_workbook

wb = load_workbook("Book1.xlsx")

sheet1 = wb.worksheets[0]
list_dict=[]
dict={}
for row_value in range(1,sheet1.max_row+1):
    dict.update({sheet1.cell(row=row_value,column=1).value:sheet1.cell(row=row_value, column=2).value})
    list_dict.append(dict)

print(list_dict)